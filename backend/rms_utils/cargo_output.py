from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import data_cleaners
import RiskModeller as rms
import functions as fnc
#from loc_file_schema import *
import pandas as pd
import numpy as np
import datetime
import re
import os
import requests
import json
from tinydb import TinyDB, Query, where
from getpass import getpass
pd.options.display.float_format = '{:,.2f}'.format
import math
import allAAL as aal
import pandas as pd
import requests
import RiskModeller as rms
from getpass import getpass
from datetime import date, datetime,timedelta
from openpyxl import load_workbook, drawing
import os
import gc

import RMmodule2 as rm2
def getLayer(row):
    return str(int(row['partOf']/1000000))+'xs'+str(int(row['attachmentPoint']/1000000))
def getPortfolioAccounts(datasource, portfolioid, bearerToken):

    ## portacct for a specific portfolioid .. select * from portacct where portinfoid = portfolioid

    path = 'portfolios/{portfolioid}/accounts?datasource=' + datasource+'&limit=100000000'
    path = path.replace('{portfolioid}',str(portfolioid))
    request = aal.sendRequest("GET",path,bearerToken).json()
    if int(request['searchTotalMatch'])>1000:
        rangeMax=int(request['searchTotalMatch']/1000+1)
        dfList=[]
        for i in range(0,rangeMax):
            offset=i*1000
            path = 'portfolios/{portfolioid}/accounts?datasource=' + datasource+'&limit=100000000&offset='+str(offset)
            path = path.replace('{portfolioid}',str(portfolioid))
            request = aal.sendRequest("GET",path,bearerToken).json()
            df_portacc = pd.DataFrame(request['searchItems'])
            dfList.append(df_portacc)
        return pd.concat(dfList).reset_index()
    else:
        df_portacc = pd.DataFrame(request['searchItems'])
        return df_portacc
def getAALmaxLocations(analysisId, perspective, refreshtoken):
    #rangeMax=22
    bearToken=refreshTokenRMS(refreshtoken)
    dfList=[]
    #df=getTopNLocationAALs(analysisId, perspective, N, bearerToken)
    path_with_param = 'analyses/{analysisid}/location-aal?perspective={perspective}&limit=1'
    path = path_with_param
    path = path.replace('{analysisid}',str(analysisId))
    path = path.replace('{perspective}',perspective)
    request = rms.sendRequest("GET",path,bearToken).json()
    print(analysisId)
    print(request)
    #return request
    rangeMax=int(request['totalCount']/1000)+1
    df_TopNLocationAALs = pd.DataFrame(request['locationAALs']) 
    for i in range(0,rangeMax):
        if i % 100 == 0:
            print(i)
        bearToken=refreshTokenRMS(refreshtoken)
        offset=i*1000
        path_with_param = 'analyses/{analysisid}/location-aal?perspective={perspective}&limit={N}&offset={offset}&sort=aal%20DESC'
        path = path_with_param
        path = path.replace('{analysisid}',str(analysisId))
        path = path.replace('{perspective}',perspective)
        path = path.replace('{N}','1000')
        path = path.replace('{offset}',str(offset))
        request = rms.sendRequest("GET",path,bearToken).json()
        df_TopNLocationAALs = pd.DataFrame(request['locationAALs']) 
        dfList.append(df_TopNLocationAALs)
            
    return pd.concat(dfList).drop_duplicates().reset_index()

def authenticationRMS(username, password):
    url = "https://api-euw1.rms.com/sml/auth/v1/login/implicit"

    payload = {"tenantName": "priceforbes","username": username,"password": password}
    headers = {
      'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data = json.dumps(payload))
    if str(response.status_code)[0]=='2':
        bearTokern = 'Bearer '+response.json()['accessToken']
        authFile = bearTokern,response.json()
    #!!!!!Exception handling!!!!!!
    else:
        authFile = "Invalid credentials"
    return authFile

def refreshTokenRMS(refreshtoken):
    url = "https://api-euw1.rms.com/sml/auth/v1/login/refresh"

    payload = {"refreshToken": refreshtoken}
    headers = {'Content-Type': 'application/json'}

    response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
    if str(response.status_code)[0]=='2':
        bearTokern='Bearer '+response.json()['accessToken']
        authFile = bearTokern
    #!!!!!Exception handling!!!!!!
    else:
        authFile = "Invalid credentials"
    return authFile

def createReport(datasource,df_analyses,df_stochastic_cep,df_portfolios,df_policies,df_locations,team,auth_file):
    start = datetime.now()
    token = auth_file[0]
    refreshtoken = auth_file[1]['refreshToken']
    ## This function completes the entire process of report generation given the
    ## analysis data, the portfolio data, the policy data, the location data, and the team
    template_file = r"ModellingTemplateV2.xlsx" ## Template file is same as for RiskLink
    output_folder = r""
    df_analyses=df_analyses.sort_values(by=['name'])
    ## As deletion from blob storage isn't part of the script, I append a date time
    ## to the file name so one can generate output more than once
    now = datetime.now()
    date_time = now.strftime("%d%m%y")
    output_file_name = "reporting_outputs/"+datasource + "_Modelling_Results_" + date_time + '.xlsx'
    output_file = output_folder + output_file_name
    print('HERE')
    ## Get the policy stats table from blob storage (until API developed)
    ## And join that table to the policytable
    #df_policystats=getPolicyStatsV2(datasource,df_analyses,df_policies,auth_file)
    # print(df_policystats)
    df_policystats=pd.DataFrame()
    if not df_policystats.empty:
    #     df_policystats = pd.merge(df_policystats, df_policies,on ='id')
    #     df_policystats = df_policystats[['AAL','analysisid','number']]
    #     ## Get the layer name from the policynum
         #df_policystats['number'] = df_policystats.apply(lambda row: row['number'][find_nth(row['number'],'-',1)+1:find_nth(row['number'],'-',2)],axis=1)
    #     ## sum over the AAl, group by analysisid and layer name
         df_policystats = df_policystats.groupby(['analysisid','number']).sum().reset_index()

    print('Creating Report')
    wb = load_workbook(template_file)

    #### ANALYSES RESULTS
    perils = {}
    analysis_results = {}
    peril_mapping = {'EQ':'Earthquake', 'WS': 'Windstorm/hurricane','AC':'Accumulation','CS':'Severe Convective Storm',
'FL':'Flood','ID': 'Infectious Disease','TO':'Tornado/hail','TR':'Probabilistic Terrorism Analysis','WC':'Works compensation/human casualty',
'WT':'Winterstorm','WF':'Wild Fire','ZZ':'Unknown','Group':'Group'}

    for index, analysis in df_analyses.iterrows():

        ### Get OEP and AEP
        analysisID = analysis['id']
        peril = 'Group' if analysis['peril'] == 'YY' else analysis['peril']['code']
        print(peril)
        print(analysis['id'])
        token=refreshTokenRMS(refreshtoken)
        print('Token refreshed')

                
#        (df_AEP_GR,df_OEP_GR) = getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GR', token)
        analysis_results['GU'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GU', token)
        try:
            analysis_results['GR'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GR', token)
        except:
            a=analysis_results['GU']
            a[0]['positionValues']=0
            a[1]['positionValues']=0
            analysis_results['GR'] = a
            print(analysisID)

        ## This is for numbering the tab. Eg. EQ (1), EQ (2)
        try:
            perils[peril] += 1
        except:
            perils[peril] =1

        ## Generate new tab for this analysis
        ws = wb["Results Template"]
        ws = wb.copy_worksheet(ws)
        ws.title = analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else peril + ' ' + str(perils[peril])
        if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
            ws.sheet_properties.tabColor = 'FFFF00'

        ## Fill in core information. Eg. datsource name, DLM run, currency run, date produced.
        c = ws.cell(row = 4, column = 3)
        c.value = analysis['name']
        c = ws.cell(row = 11, column = 3)
        c.value = peril_mapping[analysis['peril']['code']]
        c = ws.cell(row = 12, column = 3)
        c.value = analysis['description']
        c = ws.cell(row = 7, column = 3)
        c.value = analysis['currency']['code']
        c = ws.cell(row = 5, column = 3)
        c.value = date.today().year

        prem_col_dict = {'GU':7,'GR':9}
        AEP_OEP_col_dict = {'GU_AEP':7, 'GU_OEP':8,'GR_AEP':9,'GR_OEP':10}
        subset_of_analysis_results = {}

        ## Fill return period loss table and AAL
        for perspective in prem_col_dict.keys():
            ## the RMS API returns a LOT of values, so need to filter for just the ones we need.
            ## The ones we need are defined by return_periods.
            return_periods = [10000,5000,1000,500,250,200,150,100,50,25,10,5,2]
            subset_of_analysis_results[perspective + '_AEP'] = analysis_results[perspective][0][analysis_results[perspective][0]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)
            subset_of_analysis_results[perspective + '_OEP'] = analysis_results[perspective][1][analysis_results[perspective][1]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)

            ## return periods
            for i in ('AEP','OEP'):

                start_row = 8
                curr_row = start_row
                for index, df_row in subset_of_analysis_results[perspective + '_' + i].iterrows():
                    c = ws.cell(row = curr_row, column = 6)
                    c.value = df_row['returnPeriods']
                    c = ws.cell(row = curr_row, column = AEP_OEP_col_dict[perspective +'_' + i])
                    c.value = float(df_row['positionValues'])
                    curr_row +=1

        ###  Premium / AALPrint()
            print('doing Pre')

            df_premium_results = aal.getAnalysisPremResults(df_analyses[df_analyses['id']==analysisID], perspective, token)
            for index, df_row in df_premium_results.iterrows():
                c = ws.cell(row = 25, column = prem_col_dict[perspective])
                c.value = df_row['purepremium']
                c = ws.cell(row = 26, column = prem_col_dict[perspective])
                c.value = df_row['totalstddev']

        ### Policy Structure
        #df_policystats=pd.DataFrame()
        #df_policystats['number']=df_policystats['layer']
        if not df_policystats.empty:
            print('DOING POLICY STRUTURE')
            if not df_policystats[df_policystats['analysisid'] == analysisID].empty:
                policy_stats = df_policystats[df_policystats['analysisid'] == analysisID]
                print(policy_stats)
                start_row = 7
                curr_row = start_row
                for index, df_row in policy_stats.iterrows():

                    c = ws.cell(row = curr_row, column = 12)
                    c.value = df_row['number']
                    c = ws.cell(row = curr_row, column = 13)
                    c.value = df_row['AAL']
                    curr_row +=1
                if ws.cell(row = start_row, column = 12).value is None:
                    ft = Font(color='FF000000')
                    for row in range(start_row-1,start_row+22):
                        c = ws.cell(row = row, column = 12)
                        c.value = None
                        c.fill = PatternFill("solid", fgColor='ffffff')
                        c.border= Border()
                        c = ws.cell(row = row, column = 13)
                        c.fill = PatternFill("solid", fgColor='ffffff')
                        c.value = None
                        c.border= Border()

        ### Top 20 locations
        print('TOPN LOCATIONS')

        df_top20LocationAALs = aal.getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'GR', 20, token)
        if df_top20LocationAALs.empty:
            df_top20LocationAALs = aal.getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'GU', 20, token)
            df_top20LocationAALs['aal']=0
            df_top20LocationAALs['standardDeviation']=0
            df_top20LocationAALs['coefficientOfVariation']=0
        aal_col_mapping_dict = {'locationId':2,'streetAddress':3,'postalCode':4,'latitude':5,'longitude':6,'state':7, 'country':8,'geoResolutionLevel':9,'tiv':10,'aal':11,'occType':12,'constructionType':13,'yearBuilt':14,'stories':15,'floorArea':16}
        start_row = 32
        cur_row = start_row
        if not df_top20LocationAALs.empty:
            df_top20LocationAALs = df_top20LocationAALs.merge(df_locations, how='inner',on='locationId')
            df_top20LocationAALs =df_top20LocationAALs.drop_duplicates(subset=['locationNumber','aal','propertyReference'])
            for index, df_row in df_top20LocationAALs.iterrows():
                for i in aal_col_mapping_dict.keys():
                    c = ws.cell(row = cur_row, column = aal_col_mapping_dict[i])
                    c.value =  df_row[i]
                cur_row += 1
            if ws.cell(row = start_row, column = 8).value is None:
                for row in range(start_row-4,start_row+28):
                    ws.row_dimensions[row].hidden = True
        if ws.cell(row = start_row, column = 8).value is None:
            for row in range(start_row-4,start_row+28):
                ws.row_dimensions[row].hidden = True
        N = 1000000000

    #df_allLocationAALs = getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'GR', N, token)
        ## AAL
        ws = wb["AAL Detailed"]
        ws = wb.copy_worksheet(ws)
        ws.title =  'AAL ' + analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else 'AAL Detailed ' + peril + ' ' + str(perils[peril])
        if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
            ws.sheet_properties.tabColor = 'FFFF00'

        ## Fill in core information. Eg. datsource name, DLM run, currency run, date produced.
        c = ws.cell(row = 4, column = 3)
        c.value = analysis['name']
        c = ws.cell(row = 11, column = 3)
        c.value = peril_mapping[analysis['peril']['code']]
        c = ws.cell(row = 12, column = 3)
        c.value = analysis['description']
        c = ws.cell(row = 7, column = 3)
        c.value = analysis['currency']['code']
        c = ws.cell(row = 5, column = 3)
        c.value = date.today().year
        
        token=refreshTokenRMS(refreshtoken)
        
        print('Doing most of the AAL heavylifting now')
        df_aal = df_locations.copy()
        df_aal = df_aal[['locationId','streetAddress', 'postalCode','state','county', 'city', 'country', 'tiv']]
        token=refreshTokenRMS(refreshtoken)
        print('refreshed token')
        print('Made a copy of df_locations and extracted the neccessary columns')
        gr_aal = getAALmaxLocations(analysisID, 'GR', refreshtoken)
        gr_aal =gr_aal.rename(columns = {'locationId':'Locations'})
        gr_aal = gr_aal[['Locations','aal']]
        token=refreshTokenRMS(refreshtoken)
        print('refreshed token')
        print('Extracted locations of those in this analysis with AALMaxlocations function')
        df_aal = gr_aal.merge(df_aal,how = 'left',left_on = 'Locations',right_on ='locationId')
        df_state = df_aal[['aal','state','tiv']].groupby(['state']).agg(['sum'])
        df_county = df_aal[['aal','county','tiv']].groupby(['county']).agg(['sum'])
        token=refreshTokenRMS(refreshtoken)
        print('refreshed token')
        print('merged and aggregated Location and TIV information')
        AAL_insert(ws,df_state,'state')
        AAL_insert(ws,df_county,'county')   
        deleteexcess(ws,20,298,3)
        
        token=refreshTokenRMS(refreshtoken)
        print('refreshed')

        del df_aal
        del gr_aal
        gc.collect()
    if df_stochastic_cep.empty == True:
        print('stochastic_cep is empty, will go onto next step')
    else:    
        for index, analysis in df_stochastic_cep.iterrows():
            
            analysisID = analysis['id']
            peril = 'Group' if analysis['peril'] == 'YY' else analysis['peril']['code']
            print(peril)
            print(analysis['id'])
            token=refreshTokenRMS(refreshtoken)
            print('Token refreshed')
    
            analysis_results['GU'] = aal.getAnalysisCEPResults(df_stochastic_cep[df_stochastic_cep['id']==analysisID], 'GU', token)
            try:
                analysis_results['GR'] = aal.getAnalysisCEPResults(df_stochastic_cep[df_stochastic_cep['id']==analysisID], 'GR', token)
            except:
                a=analysis_results['GU']
                a[0]['positionValues']=0
                a[1]['positionValues']=0
                analysis_results['GR'] = a
                print(analysisID)
            
            ## This is for numbering the tab. Eg. EQ (1), EQ (2)
            try:
                perils[peril] += 1
            except:
                perils[peril] =1
    
            ws = wb["Event CEP"]
            ws = wb.copy_worksheet(ws)
            ws.title = analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else peril + ' ' + str(perils[peril])
            if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
                ws.sheet_properties.tabColor = 'FFFF00'
            
            c = ws.cell(row = 4, column = 3)
            c.value = analysis['name']
            c = ws.cell(row = 11, column = 3)
            c.value = peril_mapping[analysis['peril']['code']]
            c = ws.cell(row = 12, column = 3)
            c.value = analysis['description']
            c = ws.cell(row = 7, column = 3)
            c.value = analysis['currency']['code']
            c = ws.cell(row = 5, column = 3)
            c.value = date.today().year
    
            subset_of_analysis_results = {}
            prem_col_dict = {'GU':7,'GR':9}
    
            AEP_OEP_col_dict = {'GU_AEP':7, 'GU_OEP':8,'GR_AEP':9,'GR_OEP':10}
            subset_of_analysis_results = {}
    
            for perspective in prem_col_dict.keys():
            ## the RMS API returns a LOT of values, so need to filter for just the ones we need.
            ## The ones we need are defined by return_periods.
                    return_periods = [10000,5000,1000,500,250,200,150,100,50,25,10,5,2]
                    subset_of_analysis_results[perspective + '_CEP'] = analysis_results[perspective][analysis_results[perspective]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)
            
                    start_row = 8
                    curr_row = start_row
                    for index, df_row in subset_of_analysis_results[perspective + '_CEP'].iterrows():
                        c = ws.cell(row = curr_row, column = 6)
                        c.value = df_row['returnPeriods']
                        c = ws.cell(row = curr_row, column = prem_col_dict[perspective])
                        c.value = float(df_row['positionValues'])
                        curr_row +=1
    
                    print('doing Pre')
                    df_premium_results = aal.getAnalysisPremResults(df_stochastic_cep[df_stochastic_cep['id']==analysisID], perspective, token)
                    for index, df_row in df_premium_results.iterrows():
                            c = ws.cell(row = 25, column = prem_col_dict[perspective])
                            c.value = df_row['purepremium']
                            c = ws.cell(row = 26, column = prem_col_dict[perspective])
                            c.value = df_row['totalstddev']

    georesolution_col_mapping_dict = {'Count':3,'TIV':4}

    for index, portfolio in df_portfolios.iterrows():

        #### GEOCODING SUMMARY
        print('Geocoding Summary')
        df_portacct = getPortfolioAccounts(datasource, portfolio['id'], token) ## What accounts (and therefore locations) belong to each portfolio?
        if not df_portacct.empty:

            acct_list = df_portacct['accountId'].to_list()
            ws = wb["Geocoding Summary"]
            ws = wb.copy_worksheet(ws)
            ws.title = "Geocoding Summary - " + str(portfolio['id'])
            start_row = 15
            end_row = 30
            cur_row = start_row

            ## Get locations in accounts
            df_acctlocs = df_locations[df_locations['accountId'].isin(acct_list)]
            df_georesolutionlevel_all = df_acctlocs[['locationId','geoResolutionLevel','tiv']]
            ## Aggregate the TIV and number of locations in each geoResolutionCode level
            df_georesolutionlevel_all = df_georesolutionlevel_all.groupby(['geoResolutionLevel']).agg(['sum', 'count'])

            c = ws.cell(row = 4, column = 2)
            c.value = portfolio['name']
            ## Output geocoding summary to the excel
            for georesLevel, df_georesolutionlevel in df_georesolutionlevel_all.iterrows():
                #if df_georesolutionlevel['tiv']['sum'] == 0:
                    #continue;
                c = ws.cell(row = cur_row, column = 2)
                c.value = georesLevel
                c = ws.cell(row = cur_row, column = georesolution_col_mapping_dict['TIV'])
                c.value = df_georesolutionlevel['tiv']['sum']
                c = ws.cell(row = cur_row, column = georesolution_col_mapping_dict['Count'])
                c.value = df_georesolutionlevel['locationId']['count']
                cur_row += 1
            if cur_row < end_row:
                deleteexcess(ws,cur_row,end_row,2)

            df_construction = df_acctlocs[['locationId','constructionType','tiv']].groupby(['constructionType']).agg(['sum', 'count'])
            df_occupancy = df_acctlocs[['locationId','occType','tiv']].groupby(['occType']).agg(['sum', 'count'])
            df_years = df_acctlocs[['locationId','yearBracket','tiv']].groupby(['yearBracket']).agg(['sum', 'count'])
            df_stories = df_acctlocs[['locationId','storiesBracket','tiv']].groupby(['storiesBracket']).agg(['sum', 'count'])
            df_areas = df_acctlocs[['locationId','areaBracket','tiv']].groupby(['areaBracket']).agg(['sum', 'count'])
            ## DATA QUALITY

            ws = wb["Data Quality"]
            ws = wb.copy_worksheet(ws)
            ws.title = "Data Quality - " + str(portfolio['id'])
            c = ws.cell(row = 4, column = 2)
            c.value = portfolio['name']

            dataquality_insert(ws,df_construction,'Construction')
            dataquality_insert(ws,df_occupancy,'Occupancy')
            dataquality_insert(ws,df_years,'Years')
            dataquality_insert(ws,df_stories,'Stories')
            dataquality_insert(ws,df_areas,'Area')
            deleteexcess(ws,16,298,2)

    #### ASSUMPTIONS

    ws = wb["Assumptions Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Assumptions"
    start_row_assumptions = 7
    end_row = 4267
    ## PERIL LOCATION ASSUMPTIONS

    df_locationAssumptions = getLocationsAssumptions(datasource, token)
    if not df_locationAssumptions.empty:
        currency = df_row['currency'].unique()[0]
        df_locationAssumptions.drop(columns = ['currency'])
        df_locationAssumptions = df_locationAssumptions.T
        cur_row = start_row_assumptions
        for index, df_row in df_locationAssumptions.iterrows():
            c = ws.cell(row = cur_row, column = 2)
            c.value = index
            c = ws.cell(row = cur_row, column = 3)
            if df_row['0'] <=1:
                c.value = str(df_row['0']*100) + '%'
            else:
                c.value = df_row['0']
            c = ws.cell(row = cur_row, column = 4)
            c.value = currency
            cur_row += 1

    ### POLICY ASSUMPTIONS
    start_row = 50
    cur_row = start_row
    assumptions_col_mapping_dict = {'number':3,'accountNumber':2,'partOf':5,'blanketLimit':6,'minDeductible':7,'attachmentPoint':8,'blanketDeductible':9,'currency':10}

    df_accounts = aal.getAccounts(datasource, token,0)
    df_accounts = df_accounts[['accountId','accountNumber']]
    if not df_accounts.empty:
        df_polacct = df_policies.merge(df_accounts, how='inner',on='accountId')
    else:
        df_polacct = df_policies
        df_polacct['accountNumber'] = df_policies['id']
    for index, df_row in df_polacct.iterrows():
        c = ws.cell(row = cur_row, column = 4)
        c.value = df_row['peril']['name']
        for i in assumptions_col_mapping_dict.keys():
            c = ws.cell(row = cur_row, column = assumptions_col_mapping_dict[i])
            if assumptions_col_mapping_dict[i] > 3 and assumptions_col_mapping_dict[i] < 10:
                if df_row[i] <=1 and df_row[i] !=0 :
                    c.value = str(df_row[i]*100) + '%'
                else:
                    c.value = df_row[i]
            elif assumptions_col_mapping_dict[i] == 10:
                c.value = df_row[i]['code']
            else:
                c.value = df_row[i]
        cur_row += 1

    ## LOCATION SPECIFIC DEDUCTIBLES
    df_locspecded = getLocationSpecificDeductibles(df_locations)
    start_row = 646
    if not df_locspecded.empty:
        cur_row = start_row
        loc_ded_col_mapping_dict = {'lossType':2,'city':4,'county':5,'state':6,'country':7,'peril':3,'deductible':8,'currency':9}
        for index, df_row in df_locspecded.iterrows():
            for i in loc_ded_col_mapping_dict.keys():
                c = ws.cell(row = cur_row, column = loc_ded_col_mapping_dict[i])
                if i =='deductible':
                    if df_row[i] <=1 and df_row[i] !=0:
                        c.value = str(df_row[i]*100) + '%'
                    else:
                        c.value = df_row[i]
                else:
                    c.value = df_row[i]
            cur_row+=1

    if ws.cell(row = start_row, column = 3).value is None:
        for row in range(start_row-2,start_row):
            ws.row_dimensions[row].hidden = True
    if ws.cell(row = start_row_assumptions, column = 3).value is None:
        for row in range(start_row_assumptions-2,start_row_assumptions):
            ws.row_dimensions[row].hidden = True

    deleteexcess(ws,start_row_assumptions,end_row+1,2)

    ### Delete pages and Save
    ws = wb["Glossary Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Glossary"
    ws = wb["Disclaimer"]
    ws = wb.copy_worksheet(ws)
#         img = drawing.image.Image('rms.png')
#         img.width = 170
#         img.height = 165
#         img.anchor = 'A1'
#         ws.add_image(img)
    ws.title ="RMS Disclaimer"

    for i in ("Results Template","Geocoding Summary", "Data Quality", "Assumptions Template","Glossary Template","Disclaimer","ALL Template","AAL Detailed", "Event CEP"):
        ws = wb[i]
        wb.remove(ws)
    wb.save(output_file) ## save locally

    return output_file

def deleteexcess(ws,start_row,end_row,test_col):

    ### This function is used in the output script to delete excess rows in tables (eg. for occupancy type)
    for row in range(start_row,end_row):
        if ws.cell(row = row, column = test_col).value is None:
            ws.row_dimensions[row].hidden = True
        if ws.cell(row = row, column = test_col).value == 'START' or ws.cell(row = row, column = test_col).value == 'END':
            ws.cell(row = row, column = test_col).value = None
def dataquality_insert(ws,df,type):
        start_row_dict = {'Construction':16, 'Occupancy':157, 'Years':256,'Stories':273,'Area':287}
        cur_row = start_row_dict[type]
        label = 2
        dataquality_col_mapping_dict = {'count':3,'tiv':4}
        unknown_row = 0
        for index, df_row in df.iterrows():
            c = ws.cell(row = cur_row, column = label)
            c.value = index
            for i in dataquality_col_mapping_dict.keys():
                c = ws.cell(row = cur_row, column = dataquality_col_mapping_dict[i])
                c.value = df_row[i]['sum'] if i == 'tiv' else df_row['locationId'][i]
            cur_row += 1
def AAL_insert(ws,df,type):
        start_row_dict = {'state':20, 'county':161}
        cur_row = start_row_dict[type]
        label = 3
        dataquality_col_mapping_dict = {'aal':4,'tiv':5}
        unknown_row = 0
        for index, df_row in df.iterrows():
            c = ws.cell(row = cur_row, column = label)
            c.value = index
            for i in dataquality_col_mapping_dict.keys():
                c = ws.cell(row = cur_row, column = dataquality_col_mapping_dict[i])
                c.value = df_row[i]['sum'] if i == 'tiv' else df_row[i]['sum']
            cur_row += 1

def getLocationSpecificDeductibles(df_locations):

    ## gets the location specific deductibles at the level where they vary

    df_locations_spec_deds = df_locations[['city','county','state','country','currency','location']]
    try:
        df_locations_spec_deds['lossType'] = df_locations_spec_deds.apply(lambda row: row['location']['coverages']['lossType'],axis=1)
        df_locations_spec_deds['deductible'] = df_locations_spec_deds.apply(lambda row: row['location']['coverages']['deductible'],axis=1)
        df_locations_spec_deds['peril'] = df_locations_spec_deds.apply(lambda row: row['location']['coverages']['peril']['name'],axis=1)

        df_locations_spec_deds = df_locations_spec_deds.drop(columns = ['location'])

        df_ded_country = df_locations_spec_deds[['lossType', 'country','peril','deductible','currency']].drop_duplicates()
        df_ded_state = df_locations_spec_deds[['lossType', 'state','country','peril','deductible','currency']].drop_duplicates()
        df_ded_county = df_locations_spec_deds[['lossType', 'county','state','country','peril','deductible','currency']].drop_duplicates()

        if True not in df_locations_spec_deds[['lossType', 'country','peril']].duplicated().tolist():
            ### They vary at country level
            df_ded_country['state'] = None
            df_ded_country['city'] = None
            df_ded_country['county'] = None
            return df_ded_country
        elif True not in df_ded_state[['LossType','State', 'CountryCode','peril']].duplicated(keep=False).tolist():
            ## They vary at state level
            df_ded_state['city'] = None
            df_ded_state['county'] = None
            return df_ded_state
        elif True not in df_ded_county[['LossType','county','State' ,'CountryCode','peril']].duplicated(keep=False).tolist():
            ## They vary at city level
            df_ded_county['city'] = None
            return df_ded_county
        else:
            ## They vary at street level
            df_ded_county['city'] = df_ded_county.apply(lambda row: _get_city(row,df_locations_spec_deds,df_ded_county[['portid','lossType','county','state' ,'country','peril']].duplicated(keep=False)),axis=1)
            df_ded_county = df_ded_county.drop_duplicates()

        return df_ded_county
    except:
        ## There are no location specfic deductibles
        return pd.DataFrame()

def getLocationsAssumptions(datasource, bearerToken):

    ## Per peril location assumptions
    path = 'locations?datasource=' + datasource +'&limit=1000000000'

    request = aal.sendRequest("GET",path,bearerToken).json()
    getLocationsAssumptions = pd.DataFrame(request['searchItems'])
    try:
        getLocationsAssumptions['EQ'] = getLocationsAssumptions.apply(lambda row: row['location']['eqDetail']['siteDeduct'],axis=1)
        getLocationsAssumptions['currency'] = getLocationsAssumptions.apply(lambda row: row['location']['currency']['code'],axis=1)
    except:
        ## If there is no EQDET entry, then pass
        pass;
    try:
        getLocationsAssumptions['WS'] = getLocationsAssumptions.apply(lambda row: row['location']['wsDetail']['siteDeduct'],axis=1)
        getLocationsAssumptions['currency'] = getLocationsAssumptions.apply(lambda row: row['location']['currency']['code'],axis=1)
    except:
        ## If there is no HUDET entry, then pass
        pass;
    try:
        getLocationsAssumptions['CS'] = getLocationsAssumptions.apply(lambda row: row['location']['toDetail']['siteDeduct'],axis=1)
        getLocationsAssumptions['currency'] = getLocationsAssumptions.apply(lambda row: row['location']['currency']['code'],axis=1)
    except:
        ## If there is no TODET entry, then pass
        pass;
    getLocationsAssumptions = getLocationsAssumptions.drop(columns = ['location'])
    getLocationsAssumptions = getLocationsAssumptions.drop(columns = ['propertyReference'])
    getLocationsAssumptions = getLocationsAssumptions.drop_duplicates()
    return getLocationsAssumptions
def getLocationsOFFSET2(datasource, bearerToken,offset):

    ## Gets TIV and location information for each location.

    path = 'locations?datasource=' + datasource +'&limit=1000000000&offset='+str(offset)

    request = aal.sendRequest("GET",path,bearerToken).json()
    return request
def getLocationsFULL(datasource, bearToken):

    dfList=[]
    maxOffset=int(getLocationsOFFSET2(datasource, bearToken,1*1000)['searchTotalMatch']/1000)+1
    maxOffset
    for i in range(0,maxOffset):
        print(i)
        try:
            dfList.append(getLocationsOFFSET2(datasource, bearToken,i*1000)['searchItems'])
        except:
            print("EXCEPT")
            #useRefresh TOken

            #dfList.append(getLocationsOFFSET2(datasource, auth_file[0],i*1000)['searchItems'])
    dfListP=[]
    for i in dfList:
        dfListP=dfListP+i
    df_locations = pd.DataFrame(dfListP)
    df_locations['accountId'] = df_locations.apply(lambda row: row['location']['property']['accountId'],axis=1)
    df_locations['locationId'] = df_locations.apply(lambda row: row['location']['property']['locationId'],axis=1)
    df_locations['streetAddress'] = df_locations.apply(lambda row: row['location']['address']['streetAddress'],axis=1)
    df_locations['postalCode'] = df_locations.apply(lambda row: row['location']['address']['postalCode'],axis=1)
    df_locations['latitude'] = df_locations.apply(lambda row: row['location']['address']['latitude'],axis=1)
    df_locations['longitude'] = df_locations.apply(lambda row: row['location']['address']['longitude'],axis=1)
    df_locations['state'] = df_locations.apply(lambda row: row['location']['address']['admin1Name'],axis=1)
    df_locations['county'] = df_locations.apply(lambda row: row['location']['address']['admin2Name'],axis=1)
    df_locations['city'] = df_locations.apply(lambda row: row['location']['address']['cityName'],axis=1)
    df_locations['country'] = df_locations.apply(lambda row: row['location']['address']['country']['name'],axis=1)
    df_locations['cresta_zone'] = df_locations.apply(lambda row: row['location']['address']['zone1'],axis=1)
    df_locations['lowres_cresta'] = df_locations.apply(lambda row: row['location']['address']['zone3Name'],axis=1)
    df_locations['tiv'] = df_locations.apply(lambda row: row['location']['tiv'],axis=1)
    df_locations['constructionType'] = df_locations.apply(lambda row: row['location']['property']['buildingClassScheme']['code'] + row['location']['property']['buildingClass']['name'],axis=1)
    df_locations['yearBuilt'] = df_locations.apply(lambda row: row['location']['property']['yearBuilt'],axis=1)
    df_locations['stories'] = df_locations.apply(lambda row: row['location']['property']['stories'],axis=1)
    df_locations['floorArea'] = df_locations.apply(lambda row: row['location']['property']['floorArea'],axis=1)
    df_locations['occType'] = df_locations.apply(lambda row: row['location']['property']['occupancyTypeScheme']['code'] + row['location']['property']['occupancyType']['name'],axis=1)
    df_locations['geoResolutionLevel'] = df_locations.apply(lambda row: row['location']['address']['geoResolutionCode']['name'],axis=1)
    df_locations['areaBracket'] = df_locations.apply(lambda row: areaBracket(row['floorArea']),axis=1)
    df_locations['yearBracket'] = df_locations.apply(lambda row: yearBracket(row['yearBuilt']),axis=1)
    df_locations['storiesBracket'] = df_locations.apply(lambda row: storiesBracket(row['stories']),axis=1)
    df_locations['currency'] = df_locations.apply(lambda row:  row['location']['currency']['code'],axis=1)
    #df_locations=df_locations.drop('location' , axis='columns')
    return df_locations
def areaBracket(floorArea):

    ## The Floor Area brackets as defined by the Excel output files sent by Neil & Sara

    if floorArea == 'Unknown':
        return 'Unknown'
    elif int(floorArea) == 0:
        return 'Unknown'
    elif 0 < int(floorArea) < 11:
        return 'Less than 11'
    elif 11 <= int(floorArea) < 1506:
        return '11 to 1,506'
    elif 1506 <= int(floorArea) <= 2507:
        return '1,506 to 2,507'
    elif 2507 < int(floorArea) <= 5005:
        return '2,508 to 5,005'
    elif 5005 < int(floorArea) < 10010:
        return '5,006 to 10,010'
    else:
        return 'Greater than 10,011'

def yearBracket(yearBuilt):

    ## The Year Built brackets as defined by the Excel output files sent by Neil & Sara

    if yearBuilt == 'Unknown':
        return 'Unknown'
    elif int(yearBuilt) < 1937:
        return 'Pre 1937'
    elif 1937 <= int(yearBuilt) < 1973:
        return '1937 to 1973'
    elif 1973 <= int(yearBuilt) < 1988:
        return '1973 to 1988'
    elif 1988 <= int(yearBuilt) < 1991:
        return '1988 to 1991'
    elif 1991 <= int(yearBuilt) < 1994:
        return '1991 to 1994'
    elif 1994 <= int(yearBuilt) < 2001:
        return '1994 to 2001'
    elif 2001 <= int(yearBuilt)< 2010:
        return '2001 to 2010'
    elif 2010 <= int(yearBuilt) < 2050:
        return '2010 onwards'
    else:
        return 'Unknown'

def storiesBracket(stories):

    ## The Stories brackets as defined by the Excel output files sent by Neil & Sara

    if stories == 'Unknown':
        return 'Unknown'
    elif int(stories) == 0:
        return 'Unknown'
    elif 0 < int(stories) <= 3:
        return '1 - 3 (Low Rise)'
    elif 3 < int(stories) <= 7:
        return '4 - 7 (Mid Rise)'
    elif 7 < int(stories) < 14:
        return '8 - 14 (High Rise)'
    else:
        return '15 + Tall'
def find_nth(haystack, needle, n):

    ## Policy layers are defined as (eg.) EQ-25x0-OTH
    ## This function is used to find the position of the first and second "-"
    ## So I can isolate the 25x0

    start = haystack.find(needle)
    while start >= 0 and n > 1:
        start = haystack.find(needle, start+len(needle))
        n -= 1
    return start
def getPolicyStatsV2(datasource,df_analyses,df_policy,auth_file):

    #auth_file=rms.authenticationRMS('minhphan@priceforbes.com', PASSWORD)
    #df_policy=rm2.getPolicies(datasource, auth_file[0])
    #Full portfolio
    #df_analyses=rm2.getAnalyses2(datasource, auth_file[0])

    analysisids = df_analyses['id'].tolist()
    perilList = df_analyses['peril'].tolist()
    DLMlist=df_analyses['description'].tolist()
    perilList = df_analyses['peril'].tolist()
    DLMlist=df_analyses['description'].tolist()
    JobIdList=df_analyses['jobId'].tolist()

    #Iniciate Downloads in RiskMOdeller
    jobListA=[]
    for analysisid in analysisids:
        a=rm2.postRDMDownloadPolicy(datasource,analysisid,auth_file[0])
        jobListA.append(a)

    #Polling for download to finish
    jobListCurr=[]

    while len(jobListA)>0:
        for a in jobListA:
            #time.sleep(1)
            print("sleep for 6 secs")
            print(a)
            print(datasource)
            #time.sleep(6)
            rms.progressBar(str(a),auth_file)
            b=rms.getWorkStatus(str(a),auth_file)
            #if 'downloadLink' in list(response_logic.json()['response'].keys()):
            jobListCurr.append(b[0].json()['summary']['downloadLink'])
                #print(len(jobList))

            print(b[0].json()['summary']['downloadLink'])
            jobListA.remove(a)
            print(jobListA)
    #Downlaod and store file
    from datetime import date
    import zipfile, os
    from zipfile import ZipFile
    todayString=date.today().strftime("%d%m%y")
    for i in jobListCurr:
        currUrl=i
        import urllib.request
        filename=datasource+ i.split("/")[6][0:6]
        rm2._get_rdm(currUrl,filename,analysisid,todayString)
    #Read to dataFrame
    listFile=rm2.getListOfFiles('./'+todayString)
    for x in listFile:
        if  (datasource in x) and ('.zip' in x):
            print(x)
            zfile = ZipFile(x)
            zfile.extractall("./"+x[:-4])

    todayString=date.today().strftime("%d%m%y")
    listFile=rm2.getListOfFiles('./'+todayString)
    listFilePolicy=[]
    for i in listFile:
        if  ('STATS/Policy' in i) and (datasource in i):
            listFilePolicy.append(i)
    df_analyses=rm2.getAnalyses2(datasource, auth_file[0])
    df_policy=rm2.getPolicies(datasource, auth_file[0])

    #Joining data to get result
    dfPolicyList=[]
    for i in listFilePolicy:
        df=pd.read_csv(i)
        df['jobId']=i.split('/')[-1].split('_')[0]
        print(i.split('/')[-1].split('_')[0])
        dfPolicyList.append(df)
    try:
        dfPolicyFull=pd.concat(dfPolicyList)[['PolicyId','AAL','TotalStdDev','jobId']]
    except:
        dfPolicyFull=pd.concat(dfPolicyList)[['PolicyId','AAL','Std','jobId']]
    dfPolicyFull['jobId']=dfPolicyFull['jobId'].astype('int')
    df_analysesReduced=df_analyses[['dataSource', 'description','id','perilName','peril','jobId','name']]
    df_policyReduced=df_policy[['number','id']]
    df_policyReduced.columns=['PolicyNumber','PolicyId']
    dfPolicyResult=dfPolicyFull.merge(df_analysesReduced, how='left',on='jobId').merge(df_policyReduced,how='left',on='PolicyId')
    dfPolicyResult['peril']=dfPolicyResult.apply(lambda row: row['peril']['code'],axis=1)

    #delete folder
    import shutil
    shutil.rmtree('./'+todayString)
    ##rename columns
    dfPolicyResult=dfPolicyResult.rename(columns={'id':'analysisid'})
    dfPolicyResult['id']=dfPolicyResult['PolicyId']
    df_policystats = pd.merge(dfPolicyResult, df_policy,on ='id')
    df_policystats['layer']=df_policystats.apply(getLayer,axis=1)
    df_policystats['number']=df_policystats['layer']
    df_policystats = df_policystats[['AAL','analysisid','number']]
    return df_policystats

def AAL_insert_reinsurance(ws,df,type):
        start_row_dict = {'cresta_zone':20, 'lowres_cresta':161}
        cur_row = start_row_dict[type]
        label = 3
        dataquality_col_mapping_dict = {'aal':4,'tiv':5}
        unknown_row = 0
        for index, df_row in df.iterrows():
            c = ws.cell(row = cur_row, column = label)
            c.value = index
            for i in dataquality_col_mapping_dict.keys():
                c = ws.cell(row = cur_row, column = dataquality_col_mapping_dict[i])
                c.value = df_row[i]['sum'] if i == 'tiv' else df_row[i]['sum']
            cur_row += 1
            
def createReport_Reinsurance(datasource,df_analyses,df_portfolios,df_policies,df_locations,team,auth_file):
    start = datetime.now()
    token = auth_file[0]
    refreshtoken = auth_file[1]['refreshToken']
    ## This function completes the entire process of report generation given the
    ## analysis data, the portfolio data, the policy data, the location data, and the team
    template_file = r"ModellingTemplateV2_reinsurance.xlsx"
 ## Template file is same as for RiskLink
    output_folder = r""
    df_analyses=df_analyses.sort_values(by=['name'])
    ## As deletion from blob storage isn't part of the script, I append a date time
    ## to the file name so one can generate output more than once
    todayString=date.today().strftime("%d%m%y")
    output_file_name = '/opt/Analytics_Data/Analytics/RMS/Outputs/'+datasource + "_Modelling_Results_" + todayString + '.xlsx'
    output_file = output_folder + output_file_name
    print('HERE')
    ## Get the policy stats table from blob storage (until API developed)
    ## And join that table to the policytable
    df_policystats=getPolicyStatsV2(datasource,df_analyses,df_policies,auth_file)
    # print(df_policystats)
    if not df_policystats.empty:
    #     df_policystats = pd.merge(df_policystats, df_policies,on ='id')
    #     df_policystats = df_policystats[['AAL','analysisid','number']]
    #     ## Get the layer name from the policynum
         #df_policystats['number'] = df_policystats.apply(lambda row: row['number'][find_nth(row['number'],'-',1)+1:find_nth(row['number'],'-',2)],axis=1)
    #     ## sum over the AAl, group by analysisid and layer name
         df_policystats = df_policystats.groupby(['analysisid','number']).sum().reset_index()

    print('Creating Report')
    wb = load_workbook(template_file)

    #### ANALYSES RESULTS
    perils = {}
    analysis_results = {}
    peril_mapping = {'EQ':'Earthquake', 'WS': 'Windstorm/hurricane','AC':'Accumulation','CS':'Severe Convective Storm',
'FL':'Flood','ID': 'Infectious Disease','TO':'Tornado/hail','TR':'Probabilistic Terrorism Analysis','WC':'Works compensation/human casualty',
'WT':'Winterstorm','WF':'Wild Fire','ZZ':'Unknown','Group':'Group'}

    for index, analysis in df_analyses.iterrows():

        ### Get OEP and AEP
        analysisID = analysis['id']
        peril = 'Group' if analysis['peril'] == 'YY' else analysis['peril']['code']
        print(peril)
        print(analysis['id'])
        token=refreshTokenRMS(refreshtoken)
        print('Token refreshed')

                
#        (df_AEP_GR,df_OEP_GR) = getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GR', token)
        analysis_results['GU'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GU', token)
        #analysis_results['GR'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GR', token)
        try:
            #analysis_results['RL'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'RL', token)
            analysis_results['RG'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'RG', token)
        except:
            a=analysis_results['GU']
            a[0]['positionValues']=0
            a[1]['positionValues']=0
            analysis_results['GR'] = a
            print(analysisID)

        ## This is for numbering the tab. Eg. EQ (1), EQ (2)
        try:
            perils[peril] += 1
        except:
            perils[peril] =1

        ## Generate new tab for this analysis
        ws = wb["Results Template"]
        ws = wb.copy_worksheet(ws)
        ws.title = analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else peril + ' ' + str(perils[peril])
        if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
            ws.sheet_properties.tabColor = 'FFFF00'

        ## Fill in core information. Eg. datsource name, DLM run, currency run, date produced.
        c = ws.cell(row = 4, column = 3)
        c.value = analysis['name']
        c = ws.cell(row = 11, column = 3)
        c.value = peril_mapping[analysis['peril']['code']]
        c = ws.cell(row = 12, column = 3)
        c.value = analysis['description']
        c = ws.cell(row = 7, column = 3)
        c.value = analysis['currency']['code']
        c = ws.cell(row = 5, column = 3)
        c.value = date.today().year

        prem_col_dict = {'GU':7,'RG':9}
        AEP_OEP_col_dict = {'GU_AEP':7, 'GU_OEP':8,'RG_AEP':9,'RG_OEP':10}
        subset_of_analysis_results = {}

        ## Fill return period loss table and AAL
        for prem_col in prem_col_dict.keys():
            ## the RMS API returns a LOT of values, so need to filter for just the ones we need.
            ## The ones we need are defined by return_periods.
            return_periods = [10000,5000,1500,1000,500,250,200,150,100,50,25,10,5,2]
            subset_of_analysis_results[prem_col + '_AEP'] = analysis_results[prem_col][0][analysis_results[prem_col][0]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)
            subset_of_analysis_results[prem_col + '_OEP'] = analysis_results[prem_col][1][analysis_results[prem_col][1]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)

            ## return periods
            for i in ('AEP','OEP'):

                start_row = 8
                curr_row = start_row
                for index, df_row in subset_of_analysis_results[prem_col + '_' + i].iterrows():
                    c = ws.cell(row = curr_row, column = 6)
                    c.value = df_row['returnPeriods']
                    c = ws.cell(row = curr_row, column = AEP_OEP_col_dict[prem_col +'_' + i])
                    c.value = float(df_row['positionValues'])
                    curr_row +=1

        ###  Premium / AALPrint()
            print('doing Pre')

            df_premium_results = aal.getAnalysisPremResults(df_analyses[df_analyses['id']==analysisID], prem_col, token)
            for index, df_row in df_premium_results.iterrows():
                c = ws.cell(row = 26, column = prem_col_dict[prem_col])
                c.value = df_row['purepremium']
                c = ws.cell(row = 27, column = prem_col_dict[prem_col])
                c.value = df_row['totalstddev']

        ### Policy Structure
        #df_policystats=pd.DataFrame()
        #df_policystats['number']=df_policystats['layer']
        if not df_policystats.empty:
            print('DOING POLICY STRUTURE')
            if not df_policystats[df_policystats['analysisid'] == analysisID].empty:
                policy_stats = df_policystats[df_policystats['analysisid'] == analysisID]
                print(policy_stats)
                start_row = 7
                curr_row = start_row
                for index, df_row in policy_stats.iterrows():

                    c = ws.cell(row = curr_row, column = 12)
                    c.value = df_row['number']
                    c = ws.cell(row = curr_row, column = 13)
                    c.value = df_row['AAL']
                    curr_row +=1
                if ws.cell(row = start_row, column = 12).value is None:
                    ft = Font(color='FF000000')
                    for row in range(start_row-1,start_row+22):
                        c = ws.cell(row = row, column = 12)
                        c.value = None
                        c.fill = PatternFill("solid", fgColor='ffffff')
                        c.border= Border()
                        c = ws.cell(row = row, column = 13)
                        c.fill = PatternFill("solid", fgColor='ffffff')
                        c.value = None
                        c.border= Border()

        ### Top 20 locations
        print('TOPN LOCATIONS')
        df_top20LocationAALs = aal.getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'RL', 20, token)
        if df_top20LocationAALs.empty:
            df_top20LocationAALs = aal.getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'GU', 20, token)
            df_top20LocationAALs['aal']=0
            df_top20LocationAALs['standardDeviation']=0
            df_top20LocationAALs['coefficientOfVariation']=0
        aal_col_mapping_dict = {'locationId':2,'streetAddress':3,'postalCode':4,'latitude':5,'longitude':6,'state':7, 'country':8,'geoResolutionLevel':9,'tiv':10,'aal':11,'occType':12,'constructionType':13,'yearBuilt':14,'stories':15,'floorArea':16}
        start_row = 33
        cur_row = start_row
        if not df_top20LocationAALs.empty:
            df_top20LocationAALs = df_top20LocationAALs.merge(df_locations, how='inner',on='locationId')
            df_top20LocationAALs =df_top20LocationAALs.drop_duplicates(subset=['locationNumber','aal','propertyReference'])
            for index, df_row in df_top20LocationAALs.iterrows():
                for i in aal_col_mapping_dict.keys():
                    c = ws.cell(row = cur_row, column = aal_col_mapping_dict[i])
                    c.value =  df_row[i]
                cur_row += 1
            if ws.cell(row = start_row, column = 8).value is None:
                for row in range(start_row-4,start_row+28):
                    ws.row_dimensions[row].hidden = True
        if ws.cell(row = start_row, column = 8).value is None:
            for row in range(start_row-4,start_row+28):
                ws.row_dimensions[row].hidden = True
        N = 1000000000

    #df_allLocationAALs = getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'GR', N, token)
        ## AAL
        ws = wb["AAL Detailed"]
        ws = wb.copy_worksheet(ws)
        ws.title =  'AAL ' + analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else 'AAL Detailed ' + peril + ' ' + str(perils[peril])
        if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
            ws.sheet_properties.tabColor = 'FFFF00'

        ## Fill in core information. Eg. datsource name, DLM run, currency run, date produced.
        c = ws.cell(row = 4, column = 3)
        c.value = analysis['name']
        c = ws.cell(row = 11, column = 3)
        c.value = peril_mapping[analysis['peril']['code']]
        c = ws.cell(row = 12, column = 3)
        c.value = analysis['description']
        c = ws.cell(row = 7, column = 3)
        c.value = analysis['currency']['code']
        c = ws.cell(row = 5, column = 3)
        c.value = date.today().year

        token=refreshTokenRMS(refreshtoken)

        print('Doing most of the AAL heavylifting now')
        df_aal = df_locations.copy()
        df_aal = df_aal[['locationId','streetAddress', 'postalCode','cresta_zone','lowres_cresta', 'city', 'country', 'tiv']]
        token=refreshTokenRMS(refreshtoken)
        print('refreshed token')
        print('Made a copy of df_locations and extracted the neccessary columns')
        gr_aal = getAALmaxLocations(analysisID, 'RL', refreshtoken)
        gr_aal =gr_aal.rename(columns = {'locationId':'Locations'})
        gr_aal = gr_aal[['Locations','aal']]
        token=refreshTokenRMS(refreshtoken)
        print('refreshed token')
        print('Extracted locations of those in this analysis with AALMaxlocations function')
        df_aal = gr_aal.merge(df_aal,how = 'left',left_on = 'Locations',right_on ='locationId')
        df_cresta = df_aal[['aal','cresta_zone','tiv']].groupby(['cresta_zone']).agg(['sum'])
        df_lowres = df_aal[['aal','lowres_cresta','tiv']].groupby(['lowres_cresta']).agg(['sum'])
        print('merged and aggregated Location and TIV information')
        AAL_insert_reinsurance(ws,df_cresta,'cresta_zone')
        AAL_insert_reinsurance(ws,df_lowres,'lowres_cresta')   
        deleteexcess(ws,20,298,3)

        del df_aal
        del gr_aal
        gc.collect()


    georesolution_col_mapping_dict = {'Count':3,'TIV':4}

    for index, portfolio in df_portfolios.iterrows():

        #### GEOCODING SUMMARY
        print('Geocoding Summary')
        df_portacct = getPortfolioAccounts(datasource, portfolio['id'], token) ## What accounts (and therefore locations) belong to each portfolio?
        if not df_portacct.empty:

            acct_list = df_portacct['accountId'].to_list()
            ws = wb["Geocoding Summary"]
            ws = wb.copy_worksheet(ws)
            ws.title = "Geocoding Summary - " + str(portfolio['id'])
            start_row = 15
            end_row = 30
            cur_row = start_row

            ## Get locations in accounts
            df_acctlocs = df_locations[df_locations['accountId'].isin(acct_list)]
            df_georesolutionlevel_all = df_acctlocs[['locationId','geoResolutionLevel','tiv']]
            ## Aggregate the TIV and number of locations in each geoResolutionCode level
            df_georesolutionlevel_all = df_georesolutionlevel_all.groupby(['geoResolutionLevel']).agg(['sum', 'count'])

            c = ws.cell(row = 4, column = 2)
            c.value = portfolio['name']
            ## Output geocoding summary to the excel
            for georesLevel, df_georesolutionlevel in df_georesolutionlevel_all.iterrows():
                #if df_georesolutionlevel['tiv']['sum'] == 0:
                    #continue;
                c = ws.cell(row = cur_row, column = 2)
                c.value = georesLevel
                c = ws.cell(row = cur_row, column = georesolution_col_mapping_dict['TIV'])
                c.value = df_georesolutionlevel['tiv']['sum']
                c = ws.cell(row = cur_row, column = georesolution_col_mapping_dict['Count'])
                c.value = df_georesolutionlevel['locationId']['count']
                cur_row += 1
            if cur_row < end_row:
                deleteexcess(ws,cur_row,end_row,2)

            df_construction = df_acctlocs[['locationId','constructionType','tiv']].groupby(['constructionType']).agg(['sum', 'count'])
            df_occupancy = df_acctlocs[['locationId','occType','tiv']].groupby(['occType']).agg(['sum', 'count'])
            df_years = df_acctlocs[['locationId','yearBracket','tiv']].groupby(['yearBracket']).agg(['sum', 'count'])
            df_stories = df_acctlocs[['locationId','storiesBracket','tiv']].groupby(['storiesBracket']).agg(['sum', 'count'])
            df_areas = df_acctlocs[['locationId','areaBracket','tiv']].groupby(['areaBracket']).agg(['sum', 'count'])
            ## DATA QUALITY

            ws = wb["Data Quality"]
            ws = wb.copy_worksheet(ws)
            ws.title = "Data Quality - " + str(portfolio['id'])
            c = ws.cell(row = 4, column = 2)
            c.value = portfolio['name']

            dataquality_insert(ws,df_construction,'Construction')
            dataquality_insert(ws,df_occupancy,'Occupancy')
            dataquality_insert(ws,df_years,'Years')
            dataquality_insert(ws,df_stories,'Stories')
            dataquality_insert(ws,df_areas,'Area')
            deleteexcess(ws,16,298,2)

    #### ASSUMPTIONS

    ws = wb["Assumptions Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Assumptions"
    start_row_assumptions = 7
    end_row = 4267
    ## PERIL LOCATION ASSUMPTIONS

    df_locationAssumptions = getLocationsAssumptions(datasource, token)
    if not df_locationAssumptions.empty:
        currency = df_row['currency'].unique()[0]
        df_locationAssumptions.drop(columns = ['currency'])
        df_locationAssumptions = df_locationAssumptions.T
        cur_row = start_row_assumptions
        for index, df_row in df_locationAssumptions.iterrows():
            c = ws.cell(row = cur_row, column = 2)
            c.value = index
            c = ws.cell(row = cur_row, column = 3)
            if df_row['0'] <=1:
                c.value = str(df_row['0']*100) + '%'
            else:
                c.value = df_row['0']
            c = ws.cell(row = cur_row, column = 4)
            c.value = currency
            cur_row += 1

    ### POLICY ASSUMPTIONS
    start_row = 50
    cur_row = start_row
    assumptions_col_mapping_dict = {'number':3,'accountNumber':2,'partOf':5,'blanketLimit':6,'minDeductible':7,'attachmentPoint':8,'blanketDeductible':9,'currency':10}

    df_accounts = aal.getAccounts(datasource, token,0)
    df_accounts = df_accounts[['accountId','accountNumber']]
    if not df_accounts.empty:
        df_polacct = df_policies.merge(df_accounts, how='inner',on='accountId')
    else:
        df_polacct = df_policies
        df_polacct['accountNumber'] = df_policies['id']
    for index, df_row in df_polacct.iterrows():
        c = ws.cell(row = cur_row, column = 4)
        c.value = df_row['peril']['name']
        for i in assumptions_col_mapping_dict.keys():
            c = ws.cell(row = cur_row, column = assumptions_col_mapping_dict[i])
            if assumptions_col_mapping_dict[i] > 3 and assumptions_col_mapping_dict[i] < 10:
                if df_row[i] <=1 and df_row[i] !=0 :
                    c.value = str(df_row[i]*100) + '%'
                else:
                    c.value = df_row[i]
            elif assumptions_col_mapping_dict[i] == 10:
                c.value = df_row[i]['code']
            else:
                c.value = df_row[i]
        cur_row += 1

    ## LOCATION SPECIFIC DEDUCTIBLES
    df_locspecded = getLocationSpecificDeductibles(df_locations)
    start_row = 646
    if not df_locspecded.empty:
        cur_row = start_row
        loc_ded_col_mapping_dict = {'lossType':2,'city':4,'county':5,'state':6,'country':7,'peril':3,'deductible':8,'currency':9}
        for index, df_row in df_locspecded.iterrows():
            for i in loc_ded_col_mapping_dict.keys():
                c = ws.cell(row = cur_row, column = loc_ded_col_mapping_dict[i])
                if i =='deductible':
                    if df_row[i] <=1 and df_row[i] !=0:
                        c.value = str(df_row[i]*100) + '%'
                    else:
                        c.value = df_row[i]
                else:
                    c.value = df_row[i]
            cur_row+=1

    if ws.cell(row = start_row, column = 3).value is None:
        for row in range(start_row-2,start_row):
            ws.row_dimensions[row].hidden = True
    if ws.cell(row = start_row_assumptions, column = 3).value is None:
        for row in range(start_row_assumptions-2,start_row_assumptions):
            ws.row_dimensions[row].hidden = True

    deleteexcess(ws,start_row_assumptions,end_row+1,2)

    ### Delete pages and Save
    ws = wb["Glossary Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Glossary"
    ws = wb["Disclaimer"]
    ws = wb.copy_worksheet(ws)
#         img = drawing.image.Image('rms.png')
#         img.width = 170
#         img.height = 165
#         img.anchor = 'A1'
#         ws.add_image(img)
    ws.title ="RMS Disclaimer"

    for i in ("Results Template","Geocoding Summary", "Data Quality", "Assumptions Template","Glossary Template","Disclaimer","ALL Template","AAL Detailed"):
        ws = wb[i]
        wb.remove(ws)
    wb.save(output_file) ## save locally

    return output_file      

def createReport_Big(datasource,df_analyses,df_portfolios,df_policies,df_locations,team,auth_file, gr_aal):
    start = datetime.now()
    token = auth_file[0]
    refreshtoken = auth_file[1]['refreshToken']
    ## This function completes the entire process of report generation given the
    ## analysis data, the portfolio data, the policy data, the location data, and the team
    template_file = r"ModellingTemplateV2.xlsx" ## Template file is same as for RiskLink
    output_folder = r""
    df_analyses=df_analyses.sort_values(by=['name'])
    ## As deletion from blob storage isn't part of the script, I append a date time
    ## to the file name so one can generate output more than once
    now = datetime.now()
    date_time = now.strftime("%M_%S")
    output_file_name = "./outputs/"+datasource + "_Modelling_Results_" + date_time + '.xlsx'
    output_file = output_folder + output_file_name
    print('HERE')
    ## Get the policy stats table from blob storage (until API developed)
    ## And join that table to the policytable
    df_policystats=getPolicyStatsV2(datasource,df_analyses,df_policies,auth_file)
    # print(df_policystats)
    if not df_policystats.empty:
    #     df_policystats = pd.merge(df_policystats, df_policies,on ='id')
    #     df_policystats = df_policystats[['AAL','analysisid','number']]
    #     ## Get the layer name from the policynum
         #df_policystats['number'] = df_policystats.apply(lambda row: row['number'][find_nth(row['number'],'-',1)+1:find_nth(row['number'],'-',2)],axis=1)
    #     ## sum over the AAl, group by analysisid and layer name
         df_policystats = df_policystats.groupby(['analysisid','number']).sum().reset_index()

    print('Creating Report')
    wb = load_workbook(template_file)

    #### ANALYSES RESULTS
    perils = {}
    analysis_results = {}
    peril_mapping = {'EQ':'Earthquake', 'WS': 'Windstorm/hurricane','AC':'Accumulation','CS':'Severe Convective Storm',
'FL':'Flood','ID': 'Infectious Disease','TO':'Tornado/hail','TR':'Probabilistic Terrorism Analysis','WC':'Works compensation/human casualty',
'WT':'Winterstorm','WF':'Wild Fire','ZZ':'Unknown','Group':'Group'}

    for index, analysis in df_analyses.iterrows():

        ### Get OEP and AEP
        analysisID = analysis['id']
        peril = 'Group' if analysis['peril'] == 'YY' else analysis['peril']['code']
        print(peril)
        token=refreshTokenRMS(refreshtoken)
        print('Token refreshed')

                
#        (df_AEP_GR,df_OEP_GR) = getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GR', token)
        analysis_results['GU'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GU', token)
        try:
            analysis_results['GR'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GR', token)
        except:
            a=analysis_results['GU']
            a[0]['positionValues']=0
            a[1]['positionValues']=0
            analysis_results['GR'] = a
            print(analysisID)

        ## This is for numbering the tab. Eg. EQ (1), EQ (2)
        try:
            perils[peril] += 1
        except:
            perils[peril] =1

        ## Generate new tab for this analysis
        ws = wb["Results Template"]
        ws = wb.copy_worksheet(ws)
        ws.title = analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else peril + ' ' + str(perils[peril])
        if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
            ws.sheet_properties.tabColor = 'FFFF00'

        ## Fill in core information. Eg. datsource name, DLM run, currency run, date produced.
        c = ws.cell(row = 4, column = 3)
        c.value = analysis['name']
        c = ws.cell(row = 11, column = 3)
        c.value = peril_mapping[analysis['peril']['code']]
        c = ws.cell(row = 12, column = 3)
        c.value = analysis['description']
        c = ws.cell(row = 7, column = 3)
        c.value = analysis['currency']['code']
        c = ws.cell(row = 5, column = 3)
        c.value = date.today().year

        prem_col_dict = {'GU':7,'GR':9}
        AEP_OEP_col_dict = {'GU_AEP':7, 'GU_OEP':8,'GR_AEP':9,'GR_OEP':10}
        subset_of_analysis_results = {}

        ## Fill return period loss table and AAL
        for perspective in prem_col_dict.keys():
            ## the RMS API returns a LOT of values, so need to filter for just the ones we need.
            ## The ones we need are defined by return_periods.
            return_periods = [10000,5000,1000,500,250,200,150,100,50,25,10,5,2]
            subset_of_analysis_results[perspective + '_AEP'] = analysis_results[perspective][0][analysis_results[perspective][0]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)
            subset_of_analysis_results[perspective + '_OEP'] = analysis_results[perspective][1][analysis_results[perspective][1]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)

            ## return periods
            for i in ('AEP','OEP'):

                start_row = 8
                curr_row = start_row
                for index, df_row in subset_of_analysis_results[perspective + '_' + i].iterrows():
                    c = ws.cell(row = curr_row, column = 6)
                    c.value = df_row['returnPeriods']
                    c = ws.cell(row = curr_row, column = AEP_OEP_col_dict[perspective +'_' + i])
                    c.value = float(df_row['positionValues'])
                    curr_row +=1

        ###  Premium / AALPrint()
            print('doing Pre')

            df_premium_results = aal.getAnalysisPremResults(df_analyses[df_analyses['id']==analysisID], perspective, token)
            for index, df_row in df_premium_results.iterrows():
                c = ws.cell(row = 25, column = prem_col_dict[perspective])
                c.value = df_row['purepremium']
                c = ws.cell(row = 26, column = prem_col_dict[perspective])
                c.value = df_row['totalstddev']

        ### Policy Structure
        #df_policystats=pd.DataFrame()
        #df_policystats['number']=df_policystats['layer']
        if not df_policystats.empty:
            print('DOING POLICY STRUTURE')
            if not df_policystats[df_policystats['analysisid'] == analysisID].empty:
                policy_stats = df_policystats[df_policystats['analysisid'] == analysisID]
                print(policy_stats)
                start_row = 7
                curr_row = start_row
                for index, df_row in policy_stats.iterrows():

                    c = ws.cell(row = curr_row, column = 12)
                    c.value = df_row['number']
                    c = ws.cell(row = curr_row, column = 13)
                    c.value = df_row['AAL']
                    curr_row +=1
                if ws.cell(row = start_row, column = 12).value is None:
                    ft = Font(color='FF000000')
                    for row in range(start_row-1,start_row+22):
                        c = ws.cell(row = row, column = 12)
                        c.value = None
                        c.fill = PatternFill("solid", fgColor='ffffff')
                        c.border= Border()
                        c = ws.cell(row = row, column = 13)
                        c.fill = PatternFill("solid", fgColor='ffffff')
                        c.value = None
                        c.border= Border()

        ### Top 20 locations
        print('TOPN LOCATIONS')

        df_top20LocationAALs = aal.getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'GR', 20, token)
        if df_top20LocationAALs.empty:
            df_top20LocationAALs = aal.getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'GU', 20, token)
            df_top20LocationAALs['aal']=0
            df_top20LocationAALs['standardDeviation']=0
            df_top20LocationAALs['coefficientOfVariation']=0
        aal_col_mapping_dict = {'locationId':2,'streetAddress':3,'postalCode':4,'latitude':5,'longitude':6,'state':7, 'country':8,'geoResolutionLevel':9,'tiv':10,'aal':11,'occType':12,'constructionType':13,'yearBuilt':14,'stories':15,'floorArea':16}
        start_row = 32
        cur_row = start_row
        if not df_top20LocationAALs.empty:
            df_top20LocationAALs = df_top20LocationAALs.merge(df_locations, how='inner',on='locationId')
            df_top20LocationAALs =df_top20LocationAALs.drop_duplicates(subset=['locationNumber','aal','propertyReference'])
            for index, df_row in df_top20LocationAALs.iterrows():
                for i in aal_col_mapping_dict.keys():
                    c = ws.cell(row = cur_row, column = aal_col_mapping_dict[i])
                    c.value =  df_row[i]
                cur_row += 1
            if ws.cell(row = start_row, column = 8).value is None:
                for row in range(start_row-4,start_row+28):
                    ws.row_dimensions[row].hidden = True
        if ws.cell(row = start_row, column = 8).value is None:
            for row in range(start_row-4,start_row+28):
                ws.row_dimensions[row].hidden = True
        N = 1000000000

    #df_allLocationAALs = getTopNLocationAALs(df_analyses[df_analyses['id']==analysisID], 'GR', N, token)
        ## AAL
        ws = wb["AAL Detailed"]
        ws = wb.copy_worksheet(ws)
        ws.title =  'AAL ' + analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else 'AAL Detailed ' + peril + ' ' + str(perils[peril])
        if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
            ws.sheet_properties.tabColor = 'FFFF00'

        ## Fill in core information. Eg. datsource name, DLM run, currency run, date produced.
        c = ws.cell(row = 4, column = 3)
        c.value = analysis['name']
        c = ws.cell(row = 11, column = 3)
        c.value = peril_mapping[analysis['peril']['code']]
        c = ws.cell(row = 12, column = 3)
        c.value = analysis['description']
        c = ws.cell(row = 7, column = 3)
        c.value = analysis['currency']['code']
        c = ws.cell(row = 5, column = 3)
        c.value = date.today().year
        
        token=refreshTokenRMS(refreshtoken)
        
        print('Doing most of the AAL heavylifting now')
        df_aal = df_locations.copy()
        df_aal = df_aal[['locationId','streetAddress', 'postalCode','state','county', 'city', 'country', 'tiv','cresta_zone','lowres_cresta']]
        token=refreshTokenRMS(refreshtoken)
        print('refreshed token')
        print('Made a copy of df_locations and extracted the neccessary columns')
        gr_aal2 = gr_aal[gr_aal['analysisid'] == analysisID]
        gr_aal2 =gr_aal2.rename(columns = {'locationId':'Locations'})
        gr_aal2 = gr_aal2[['Locations','aal']]
        token=refreshTokenRMS(refreshtoken)
        print('refreshed token')
        print('Extracted locations of those in this analysis with AALMaxlocations function')
        df_aal = gr_aal2.merge(df_aal,how = 'left',left_on = 'Locations',right_on ='locationId')
        #df_state = df_aal[['aal','state','tiv']].groupby(['state']).agg(['sum'])
        #df_county = df_aal[['aal','county','tiv']].groupby(['county']).agg(['sum'])
        df_cresta = df_aal[['aal','cresta_zone','tiv']].groupby(['cresta_zone']).agg(['sum'])
        df_lowres = df_aal[['aal','lowres_cresta','tiv']].groupby(['lowres_cresta']).agg(['sum'])
        token=refreshTokenRMS(refreshtoken)
        print('refreshed token')
        print('merged and aggregated Location and TIV information')
        AAL_insert(ws,df_cresta,'state')
        AAL_insert(ws,df_lowres,'county') 
        deleteexcess(ws,20,298,3)
        
        token=refreshTokenRMS(refreshtoken)
        print('refreshed')

        del df_aal
        del gr_aal2
        gc.collect()


    georesolution_col_mapping_dict = {'Count':3,'TIV':4}

    for index, portfolio in df_portfolios.iterrows():

        #### GEOCODING SUMMARY
        print('Geocoding Summary')
        df_portacct = getPortfolioAccounts(datasource, portfolio['id'], token) ## What accounts (and therefore locations) belong to each portfolio?
        if not df_portacct.empty:

            acct_list = df_portacct['accountId'].to_list()
            ws = wb["Geocoding Summary"]
            ws = wb.copy_worksheet(ws)
            ws.title = "Geocoding Summary - " + str(portfolio['id'])
            start_row = 15
            end_row = 30
            cur_row = start_row

            ## Get locations in accounts
            df_acctlocs = df_locations[df_locations['accountId'].isin(acct_list)]
            df_georesolutionlevel_all = df_acctlocs[['locationId','geoResolutionLevel','tiv']]
            ## Aggregate the TIV and number of locations in each geoResolutionCode level
            df_georesolutionlevel_all = df_georesolutionlevel_all.groupby(['geoResolutionLevel']).agg(['sum', 'count'])

            c = ws.cell(row = 4, column = 2)
            c.value = portfolio['name']
            ## Output geocoding summary to the excel
            for georesLevel, df_georesolutionlevel in df_georesolutionlevel_all.iterrows():
                #if df_georesolutionlevel['tiv']['sum'] == 0:
                    #continue;
                c = ws.cell(row = cur_row, column = 2)
                c.value = georesLevel
                c = ws.cell(row = cur_row, column = georesolution_col_mapping_dict['TIV'])
                c.value = df_georesolutionlevel['tiv']['sum']
                c = ws.cell(row = cur_row, column = georesolution_col_mapping_dict['Count'])
                c.value = df_georesolutionlevel['locationId']['count']
                cur_row += 1
            if cur_row < end_row:
                deleteexcess(ws,cur_row,end_row,2)

            df_construction = df_acctlocs[['locationId','constructionType','tiv']].groupby(['constructionType']).agg(['sum', 'count'])
            df_occupancy = df_acctlocs[['locationId','occType','tiv']].groupby(['occType']).agg(['sum', 'count'])
            df_years = df_acctlocs[['locationId','yearBracket','tiv']].groupby(['yearBracket']).agg(['sum', 'count'])
            df_stories = df_acctlocs[['locationId','storiesBracket','tiv']].groupby(['storiesBracket']).agg(['sum', 'count'])
            df_areas = df_acctlocs[['locationId','areaBracket','tiv']].groupby(['areaBracket']).agg(['sum', 'count'])
            ## DATA QUALITY

            ws = wb["Data Quality"]
            ws = wb.copy_worksheet(ws)
            ws.title = "Data Quality - " + str(portfolio['id'])
            c = ws.cell(row = 4, column = 2)
            c.value = portfolio['name']

            dataquality_insert(ws,df_construction,'Construction')
            dataquality_insert(ws,df_occupancy,'Occupancy')
            dataquality_insert(ws,df_years,'Years')
            dataquality_insert(ws,df_stories,'Stories')
            dataquality_insert(ws,df_areas,'Area')
            deleteexcess(ws,16,298,2)

    #### ASSUMPTIONS

    ws = wb["Assumptions Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Assumptions"
    start_row_assumptions = 7
    end_row = 4267
    ## PERIL LOCATION ASSUMPTIONS

    df_locationAssumptions = getLocationsAssumptions(datasource, token)
    if not df_locationAssumptions.empty:
        currency = df_row['currency'].unique()[0]
        df_locationAssumptions.drop(columns = ['currency'])
        df_locationAssumptions = df_locationAssumptions.T
        cur_row = start_row_assumptions
        for index, df_row in df_locationAssumptions.iterrows():
            c = ws.cell(row = cur_row, column = 2)
            c.value = index
            c = ws.cell(row = cur_row, column = 3)
            if df_row['0'] <=1:
                c.value = str(df_row['0']*100) + '%'
            else:
                c.value = df_row['0']
            c = ws.cell(row = cur_row, column = 4)
            c.value = currency
            cur_row += 1

    ### POLICY ASSUMPTIONS
    start_row = 50
    cur_row = start_row
    assumptions_col_mapping_dict = {'number':3,'accountNumber':2,'partOf':5,'blanketLimit':6,'minDeductible':7,'attachmentPoint':8,'blanketDeductible':9,'currency':10}

    df_accounts = aal.getAccounts(datasource, token,0)
    df_accounts = df_accounts[['accountId','accountNumber']]
    if not df_accounts.empty:
        df_polacct = df_policies.merge(df_accounts, how='inner',on='accountId')
    else:
        df_polacct = df_policies
        df_polacct['accountNumber'] = df_policies['id']
    for index, df_row in df_polacct.iterrows():
        c = ws.cell(row = cur_row, column = 4)
        c.value = df_row['peril']['name']
        for i in assumptions_col_mapping_dict.keys():
            c = ws.cell(row = cur_row, column = assumptions_col_mapping_dict[i])
            if assumptions_col_mapping_dict[i] > 3 and assumptions_col_mapping_dict[i] < 10:
                if df_row[i] <=1 and df_row[i] !=0 :
                    c.value = str(df_row[i]*100) + '%'
                else:
                    c.value = df_row[i]
            elif assumptions_col_mapping_dict[i] == 10:
                c.value = df_row[i]['code']
            else:
                c.value = df_row[i]
        cur_row += 1

    ##### LOCATION SPECIFIC DEDUCTIBLES
    ###df_locspecded = getLocationSpecificDeductibles(df_locations)
    ###start_row = 646
    ###if not df_locspecded.empty:
    ###    cur_row = start_row
    ###    loc_ded_col_mapping_dict = {'lossType':2,'city':4,'county':5,'state':6,'country':7,'peril':3,'deductible':8,'currency':9}
    ###    for index, df_row in df_locspecded.iterrows():
    ###        for i in loc_ded_col_mapping_dict.keys():
    ###            c = ws.cell(row = cur_row, column = loc_ded_col_mapping_dict[i])
    ###            if i =='deductible':
    ###                if df_row[i] <=1 and df_row[i] !=0:
    ###                    c.value = str(df_row[i]*100) + '%'
    ###                else:
    ###                    c.value = df_row[i]
    ###            else:
    ###                c.value = df_row[i]
    ###        cur_row+=1

    ###if ws.cell(row = start_row, column = 3).value is None:
    ###    for row in range(start_row-2,start_row):
    ###        ws.row_dimensions[row].hidden = True
    ###if ws.cell(row = start_row_assumptions, column = 3).value is None:
    ###    for row in range(start_row_assumptions-2,start_row_assumptions):
    ###        ws.row_dimensions[row].hidden = True
    deleteexcess(ws,start_row_assumptions,end_row+1,2)

    ### Delete pages and Save
    ws = wb["Glossary Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Glossary"
    ws = wb["Disclaimer"]
    ws = wb.copy_worksheet(ws)
#         img = drawing.image.Image('rms.png')
#         img.width = 170
#         img.height = 165
#         img.anchor = 'A1'
#         ws.add_image(img)
    ws.title ="RMS Disclaimer"

    for i in ("Results Template","Geocoding Summary", "Data Quality", "Assumptions Template","Glossary Template","Disclaimer","ALL Template","AAL Detailed"):
        ws = wb[i]
        wb.remove(ws)
    wb.save(output_file) ## save locally

    return output_file


def createReport_condensed(datasource,df_analyses,df_portfolios,df_policies,df_locations,team,auth_file):
    start = datetime.now()
    token = auth_file[0]
    refreshtoken = auth_file[1]['refreshToken']
    ## This function completes the entire process of report generation given the
    ## analysis data, the portfolio data, the policy data, the location data, and the team
    template_file = r"ModellingTemplateV2.xlsx" ## Template file is same as for RiskLink
    output_folder = r""
    df_analyses=df_analyses.sort_values(by=['name'])
    ## As deletion from blob storage isn't part of the script, I append a date time
    ## to the file name so one can generate output more than once
    now = datetime.now()
    date_time = now.strftime("%M_%S")
    output_file_name = "./outputs/"+datasource + "_Modelling_Results_" + date_time + '.xlsx'
    output_file = output_folder + output_file_name
    print('HERE')
    ## Get the policy stats table from blob storage (until API developed)
    ## And join that table to the policytable
    df_policystats=getPolicyStatsV2(datasource,df_analyses,df_policies,auth_file)
    # print(df_policystats)
    if not df_policystats.empty:
    #     df_policystats = pd.merge(df_policystats, df_policies,on ='id')
    #     df_policystats = df_policystats[['AAL','analysisid','number']]
    #     ## Get the layer name from the policynum
         #df_policystats['number'] = df_policystats.apply(lambda row: row['number'][find_nth(row['number'],'-',1)+1:find_nth(row['number'],'-',2)],axis=1)
    #     ## sum over the AAl, group by analysisid and layer name
         df_policystats = df_policystats.groupby(['analysisid','number']).sum().reset_index()

    print('Creating Report')
    wb = load_workbook(template_file)

    #### ANALYSES RESULTS
    perils = {}
    analysis_results = {}
    peril_mapping = {'EQ':'Earthquake', 'WS': 'Windstorm/hurricane','AC':'Accumulation','CS':'Severe Convective Storm',
'FL':'Flood','ID': 'Infectious Disease','TO':'Tornado/hail','TR':'Probabilistic Terrorism Analysis','WC':'Works compensation/human casualty',
'WT':'Winterstorm','WF':'Wild Fire','ZZ':'Unknown','Group':'Group'}

    for index, analysis in df_analyses.iterrows():

        ### Get OEP and AEP
        analysisID = analysis['id']
        peril = 'Group' if analysis['peril'] == 'YY' else analysis['peril']['code']
        print(peril)
        token=refreshTokenRMS(refreshtoken)
        print('Token refreshed')

                
#        (df_AEP_GR,df_OEP_GR) = getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GR', token)
        analysis_results['GU'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GU', token)
        try:
            analysis_results['GR'] = aal.getAnalysisEPResults(df_analyses[df_analyses['id']==analysisID], 'GR', token)
        except:
            a=analysis_results['GU']
            a[0]['positionValues']=0
            a[1]['positionValues']=0
            analysis_results['GR'] = a
            print(analysisID)

        ## This is for numbering the tab. Eg. EQ (1), EQ (2)
        try:
            perils[peril] += 1
        except:
            perils[peril] =1

        ## Generate new tab for this analysis
        ws = wb["Results Template"]
        ws = wb.copy_worksheet(ws)
        ws.title = analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else peril + ' ' + str(perils[peril])
        if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
            ws.sheet_properties.tabColor = 'FFFF00'

        ## Fill in core information. Eg. datsource name, DLM run, currency run, date produced.
        c = ws.cell(row = 4, column = 3)
        c.value = analysis['name']
        c = ws.cell(row = 11, column = 3)
        c.value = peril_mapping[analysis['peril']['code']]
        c = ws.cell(row = 12, column = 3)
        c.value = analysis['description']
        c = ws.cell(row = 7, column = 3)
        c.value = analysis['currency']['code']
        c = ws.cell(row = 5, column = 3)
        c.value = date.today().year

        prem_col_dict = {'GU':7,'GR':9}
        AEP_OEP_col_dict = {'GU_AEP':7, 'GU_OEP':8,'GR_AEP':9,'GR_OEP':10}
        subset_of_analysis_results = {}

        ## Fill return period loss table and AAL
        for perspective in prem_col_dict.keys():
            ## the RMS API returns a LOT of values, so need to filter for just the ones we need.
            ## The ones we need are defined by return_periods.
            return_periods = [10000,5000,1000,500,250,200,150,100,50,25,10,5,2]
            subset_of_analysis_results[perspective + '_AEP'] = analysis_results[perspective][0][analysis_results[perspective][0]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)
            subset_of_analysis_results[perspective + '_OEP'] = analysis_results[perspective][1][analysis_results[perspective][1]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)

            ## return periods
            for i in ('AEP','OEP'):

                start_row = 8
                curr_row = start_row
                for index, df_row in subset_of_analysis_results[perspective + '_' + i].iterrows():
                    c = ws.cell(row = curr_row, column = 6)
                    c.value = df_row['returnPeriods']
                    c = ws.cell(row = curr_row, column = AEP_OEP_col_dict[perspective +'_' + i])
                    c.value = float(df_row['positionValues'])
                    curr_row +=1

        ###  Premium / AALPrint()
            print('doing Pre')

            df_premium_results = aal.getAnalysisPremResults(df_analyses[df_analyses['id']==analysisID], perspective, token)
            for index, df_row in df_premium_results.iterrows():
                c = ws.cell(row = 25, column = prem_col_dict[perspective])
                c.value = df_row['purepremium']
                c = ws.cell(row = 26, column = prem_col_dict[perspective])
                c.value = df_row['totalstddev']

        ### Policy Structure
        #df_policystats=pd.DataFrame()
        #df_policystats['number']=df_policystats['layer']
        if not df_policystats.empty:
            print('DOING POLICY STRUTURE')
            if not df_policystats[df_policystats['analysisid'] == analysisID].empty:
                policy_stats = df_policystats[df_policystats['analysisid'] == analysisID]
                print(policy_stats)
                start_row = 7
                curr_row = start_row
                for index, df_row in policy_stats.iterrows():

                    c = ws.cell(row = curr_row, column = 12)
                    c.value = df_row['number']
                    c = ws.cell(row = curr_row, column = 13)
                    c.value = df_row['AAL']
                    curr_row +=1
                if ws.cell(row = start_row, column = 12).value is None:
                    ft = Font(color='FF000000')
                    for row in range(start_row-1,start_row+22):
                        c = ws.cell(row = row, column = 12)
                        c.value = None
                        c.fill = PatternFill("solid", fgColor='ffffff')
                        c.border= Border()
                        c = ws.cell(row = row, column = 13)
                        c.fill = PatternFill("solid", fgColor='ffffff')
                        c.value = None
                        c.border= Border()

    georesolution_col_mapping_dict = {'Count':3,'TIV':4}

    for index, portfolio in df_portfolios.iterrows():

        #### GEOCODING SUMMARY
        print('Geocoding Summary')
        df_portacct = getPortfolioAccounts(datasource, portfolio['id'], token) ## What accounts (and therefore locations) belong to each portfolio?
        if not df_portacct.empty:

            acct_list = df_portacct['accountId'].to_list()
            ws = wb["Geocoding Summary"]
            ws = wb.copy_worksheet(ws)
            ws.title = "Geocoding Summary - " + str(portfolio['id'])
            start_row = 15
            end_row = 30
            cur_row = start_row

            ## Get locations in accounts
            df_acctlocs = df_locations[df_locations['accountId'].isin(acct_list)]
            df_georesolutionlevel_all = df_acctlocs[['locationId','geoResolutionLevel','tiv']]
            ## Aggregate the TIV and number of locations in each geoResolutionCode level
            df_georesolutionlevel_all = df_georesolutionlevel_all.groupby(['geoResolutionLevel']).agg(['sum', 'count'])

            c = ws.cell(row = 4, column = 2)
            c.value = portfolio['name']
            ## Output geocoding summary to the excel
            for georesLevel, df_georesolutionlevel in df_georesolutionlevel_all.iterrows():
                #if df_georesolutionlevel['tiv']['sum'] == 0:
                    #continue;
                c = ws.cell(row = cur_row, column = 2)
                c.value = georesLevel
                c = ws.cell(row = cur_row, column = georesolution_col_mapping_dict['TIV'])
                c.value = df_georesolutionlevel['tiv']['sum']
                c = ws.cell(row = cur_row, column = georesolution_col_mapping_dict['Count'])
                c.value = df_georesolutionlevel['locationId']['count']
                cur_row += 1
            if cur_row < end_row:
                deleteexcess(ws,cur_row,end_row,2)

            df_construction = df_acctlocs[['locationId','constructionType','tiv']].groupby(['constructionType']).agg(['sum', 'count'])
            df_occupancy = df_acctlocs[['locationId','occType','tiv']].groupby(['occType']).agg(['sum', 'count'])
            df_years = df_acctlocs[['locationId','yearBracket','tiv']].groupby(['yearBracket']).agg(['sum', 'count'])
            df_stories = df_acctlocs[['locationId','storiesBracket','tiv']].groupby(['storiesBracket']).agg(['sum', 'count'])
            df_areas = df_acctlocs[['locationId','areaBracket','tiv']].groupby(['areaBracket']).agg(['sum', 'count'])
            ## DATA QUALITY

            ws = wb["Data Quality"]
            ws = wb.copy_worksheet(ws)
            ws.title = "Data Quality - " + str(portfolio['id'])
            c = ws.cell(row = 4, column = 2)
            c.value = portfolio['name']

            dataquality_insert(ws,df_construction,'Construction')
            dataquality_insert(ws,df_occupancy,'Occupancy')
            dataquality_insert(ws,df_years,'Years')
            dataquality_insert(ws,df_stories,'Stories')
            dataquality_insert(ws,df_areas,'Area')
            deleteexcess(ws,16,298,2)

    #### ASSUMPTIONS

    ws = wb["Assumptions Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Assumptions"
    start_row_assumptions = 7
    end_row = 4267
    ## PERIL LOCATION ASSUMPTIONS

    df_locationAssumptions = getLocationsAssumptions(datasource, token)
    if not df_locationAssumptions.empty:
        currency = df_row['currency'].unique()[0]
        df_locationAssumptions.drop(columns = ['currency'])
        df_locationAssumptions = df_locationAssumptions.T
        cur_row = start_row_assumptions
        for index, df_row in df_locationAssumptions.iterrows():
            c = ws.cell(row = cur_row, column = 2)
            c.value = index
            c = ws.cell(row = cur_row, column = 3)
            if df_row['0'] <=1:
                c.value = str(df_row['0']*100) + '%'
            else:
                c.value = df_row['0']
            c = ws.cell(row = cur_row, column = 4)
            c.value = currency
            cur_row += 1

    ### POLICY ASSUMPTIONS
    start_row = 50
    cur_row = start_row
    assumptions_col_mapping_dict = {'number':3,'accountNumber':2,'partOf':5,'blanketLimit':6,'minDeductible':7,'attachmentPoint':8,'blanketDeductible':9,'currency':10}

    df_accounts = aal.getAccounts(datasource, token,0)
    df_accounts = df_accounts[['accountId','accountNumber']]
    if not df_accounts.empty:
        df_polacct = df_policies.merge(df_accounts, how='inner',on='accountId')
    else:
        df_polacct = df_policies
        df_polacct['accountNumber'] = df_policies['id']
    for index, df_row in df_polacct.iterrows():
        c = ws.cell(row = cur_row, column = 4)
        c.value = df_row['peril']['name']
        for i in assumptions_col_mapping_dict.keys():
            c = ws.cell(row = cur_row, column = assumptions_col_mapping_dict[i])
            if assumptions_col_mapping_dict[i] > 3 and assumptions_col_mapping_dict[i] < 10:
                if df_row[i] <=1 and df_row[i] !=0 :
                    c.value = str(df_row[i]*100) + '%'
                else:
                    c.value = df_row[i]
            elif assumptions_col_mapping_dict[i] == 10:
                c.value = df_row[i]['code']
            else:
                c.value = df_row[i]
        cur_row += 1

    deleteexcess(ws,start_row_assumptions,end_row+1,2)

    ### Delete pages and Save
    ws = wb["Glossary Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Glossary"
    ws = wb["Disclaimer"]
    ws = wb.copy_worksheet(ws)
#         img = drawing.image.Image('rms.png')
#         img.width = 170
#         img.height = 165
#         img.anchor = 'A1'
#         ws.add_image(img)
    ws.title ="RMS Disclaimer"

    for i in ("Results Template","Geocoding Summary", "Data Quality", "Assumptions Template","Glossary Template","Disclaimer","ALL Template","AAL Detailed"):
        ws = wb[i]
        wb.remove(ws)
    wb.save(output_file) ## save locally

    return output_file

def EventResponse_Report(datasource,df_analyses,df_stochastic_cep,df_portfolios,df_policies,df_locations,team,auth_file):
    start = datetime.now()
    token = auth_file[0]
    refreshtoken = auth_file[1]['refreshToken']
    ## This function completes the entire process of report generation given the
    ## analysis data, the portfolio data, the policy data, the location data, and the team
    template_file = r"ModellingTemplateV2.xlsx" ## Template file is same as for RiskLink
    output_folder = r""
    ## As deletion from blob storage isn't part of the script, I append a date time
    ## to the file name so one can generate output more than once
    now = datetime.now()
    date_time = now.strftime("%d%m%y")
    output_file_name = "/opt/Analytics_Data/Analytics/RMS/Outputs/"+datasource + "_Event_Response_" + date_time + '.xlsx'
    output_file = output_folder + output_file_name
    print('HERE')
    ## Get the policy stats table from blob storage (until API developed)
    ## And join that table to the policytable
##       df_policystats=getPolicyStatsV2(datasource,df_analyses,df_policies,auth_file)
##       # print(df_policystats)
##       if not df_policystats.empty:
##       #     df_policystats = pd.merge(df_policystats, df_policies,on ='id')
##       #     df_policystats = df_policystats[['AAL','analysisid','number']]
##       #     ## Get the layer name from the policynum
##            #df_policystats['number'] = df_policystats.apply(lambda row: row['number'][find_nth(row['number'],'-',1)+1:find_nth(row['number'],'-',2)],axis=1)
##       #     ## sum over the AAl, group by analysisid and layer name
##            df_policystats = df_policystats.groupby(['analysisid','number']).sum().reset_index()

    print('Creating Report')
    wb = load_workbook(template_file)

    #### ANALYSES RESULTS
    perils = {}
    analysis_results = {}
    peril_mapping = {'EQ':'Earthquake', 'WS': 'Windstorm/hurricane','AC':'Accumulation','CS':'Severe Convective Storm',
'FL':'Flood','ID': 'Infectious Disease','TO':'Tornado/hail','TR':'Probabilistic Terrorism Analysis','WC':'Works compensation/human casualty',
'WT':'Winterstorm','WF':'Wild Fire','ZZ':'Unknown','Group':'Group'}
  
    for index, analysis in df_stochastic_cep.iterrows():
        
        analysisID = analysis['id']
        peril = 'Group' if analysis['peril'] == 'YY' else analysis['peril']['code']
        print(peril)
        print(analysis['id'])
        token=refreshTokenRMS(refreshtoken)
        print('Token refreshed')

        analysis_results['GU'] = aal.getAnalysisCEPResults(df_stochastic_cep[df_stochastic_cep['id']==analysisID], 'GU', token)
        try:
            analysis_results['GR'] = aal.getAnalysisCEPResults(df_stochastic_cep[df_stochastic_cep['id']==analysisID], 'GR', token)
        except:
            a=analysis_results['GU']
            a[0]['positionValues']=0
            a[1]['positionValues']=0
            analysis_results['GR'] = a
            print(analysisID)
        
        ## This is for numbering the tab. Eg. EQ (1), EQ (2)
        try:
            perils[peril] += 1
        except:
            perils[peril] =1

        ws = wb["Event CEP"]
        ws = wb.copy_worksheet(ws)
        ws.title = analysis['name'] + ' - Group' if analysis['peril'] == 'YY' else peril + ' ' + str(perils[peril])
        if analysis['peril']['code'] == 'YY': ## If it's a grouped analysis, colour it FFFF00
            ws.sheet_properties.tabColor = 'FFFF00'
        
        c = ws.cell(row = 4, column = 3)
        c.value = analysis['name']
        c = ws.cell(row = 11, column = 3)
        c.value = peril_mapping[analysis['peril']['code']]
        c = ws.cell(row = 12, column = 3)
        c.value = analysis['description']
        c = ws.cell(row = 7, column = 3)
        c.value = analysis['currency']['code']
        c = ws.cell(row = 5, column = 3)
        c.value = date.today().year

        subset_of_analysis_results = {}
        prem_col_dict = {'GU':7,'GR':9}

        AEP_OEP_col_dict = {'GU_AEP':7, 'GU_OEP':8,'GR_AEP':9,'GR_OEP':10}
        subset_of_analysis_results = {}

        for perspective in prem_col_dict.keys():
        ## the RMS API returns a LOT of values, so need to filter for just the ones we need.
        ## The ones we need are defined by return_periods.
                return_periods = [10000,5000,1000,500,250,200,150,100,50,25,10,5,2]
                subset_of_analysis_results[perspective + '_CEP'] = analysis_results[perspective][analysis_results[perspective]['returnPeriods'].isin(return_periods)].sort_values('returnPeriods', ascending=False)
        
                start_row = 8
                curr_row = start_row
                for index, df_row in subset_of_analysis_results[perspective + '_CEP'].iterrows():
                    c = ws.cell(row = curr_row, column = 6)
                    c.value = df_row['returnPeriods']
                    c = ws.cell(row = curr_row, column = prem_col_dict[perspective])
                    c.value = float(df_row['positionValues'])
                    curr_row +=1

                print('doing Pre')
                df_premium_results = aal.getAnalysisPremResults(df_stochastic_cep[df_stochastic_cep['id']==analysisID], perspective, token)
                for index, df_row in df_premium_results.iterrows():
                        c = ws.cell(row = 25, column = prem_col_dict[perspective])
                        c.value = df_row['purepremium']
                        c = ws.cell(row = 26, column = prem_col_dict[perspective])
                        c.value = df_row['totalstddev']

    georesolution_col_mapping_dict = {'Count':3,'TIV':4}

    ### Delete pages and Save
    ws = wb["Glossary Template"]
    ws = wb.copy_worksheet(ws)
    ws.title = "Glossary"
    ws = wb["Disclaimer"]
    ws = wb.copy_worksheet(ws)
#         img = drawing.image.Image('rms.png')
#         img.width = 170
#         img.height = 165
#         img.anchor = 'A1'
#         ws.add_image(img)
    ws.title ="RMS Disclaimer"

    for i in ("Results Template","Geocoding Summary", "Data Quality", "Assumptions Template","Glossary Template","Disclaimer","ALL Template","AAL Detailed", "Event CEP"):
        ws = wb[i]
        wb.remove(ws)
    wb.save(output_file) ## save locally

    return output_file